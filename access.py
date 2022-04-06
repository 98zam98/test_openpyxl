from openpyxl import Workbook, load_workbook

wb = load_workbook("genarated.xlsx")


ws = wb.active

print(ws)


print(ws['C2'].value)





#--------------------------------------------------------------------


#  RN73

# Size : ['2B','2A','2E','1J','1E']

# Termination : ['T','L']

# Packaging :  ['TD','TDD','TED','TE','TP']
# Resistance : D2 0-9 
# ResistanceD1 : 0-9 R
# ResistanceD0 : 0-9 

# Tolerance : ['A','B','C','D','F']

# tcr :  ['05','10','25','50','100']



family = 'RN73'
Size = ['2B','2A','2E','1J','1E']

Termination = ['T','L']

Packaging =  ['TD','TDD','TED','TE','TP']
Resistance = random.randint(10, 99)
ResistanceD1 = "R" if random.randint(0, 1) else random.randint(0, 9)
ResistanceD0 = random.randint(0, 9)

Tolerance = ['A','B','C','D','F']

tcr =  ['05','10','25','50','100']


#datarandom = 'RN73'+Size[random.randint(0, 5)]+Termination[random.randint(0, 5)]+Packaging[random.randint(0, 5)]+random.randint(10, 99)+ResistanceD1+random.randint(0, 9)+Tolerance[random.randint(0, 5)]+tcr[random.randint(0, 5)]


#--------------------------------------------------------------------------
#string.startswith(value, start, end)
#rest = [ws['A1'].value[i:j]
family_flag = ws['A1'].value.startswith(family)
rest = ws['A1'].value.replace(family, "")
