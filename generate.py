from openpyxl import Workbook, load_workbook
import random



#--------------------------------------------------------------------


#  RN73

# Size : ['2B','2A','2E','1J','1E']

# Termination : ['T','L']

# Packaging :  ['TD','TDD','TED','TE','TP']
# Resistance : D2 0-9 
# ResistanceD1 : 0-9 R
# ResistanceD0 : 0-9 

# Tolerance : ['A','B','C','D','F']

# tcr :  ['05','10','25','50','100']




Size = ['2B','2A','2E','1J','1E']

Termination = ['T','L']

Packaging =  ['TD','TDD','TED','TE','TP']
Resistance = random.randint(10, 99)
ResistanceD1 = "R" if random.randint(0, 1) else random.randint(0, 9)
ResistanceD0 = random.randint(0, 9)

Tolerance = ['A','B','C','D','F']

tcr =  ['05','10','25','50','100']


#datarandom = 'RN73'+Size[random.randint(0, 5)]+Termination[random.randint(0, 5)]+Packaging[random.randint(0, 5)]+random.randint(10, 99)+ResistanceD1+random.randint(0, 9)+Tolerance[random.randint(0, 5)]+tcr[random.randint(0, 5)]

def datarandom():
    ResistanceD1 = "R" if random.randint(0, 1) else str(random.randint(0, 9))
    return 'RN73'+Size[random.randint(0, 4)]+Termination[random.randint(0, 1)]+Packaging[random.randint(0, 4)]+str(random.randint(10, 99))+ResistanceD1+str(random.randint(0, 9))+Tolerance[random.randint(0, 4)]+tcr[random.randint(0, 4)]

def rowrandom():
    ResistanceD1 = "R" if random.randint(0, 1) else str(random.randint(0, 9))

    zpart_number = 'RN73'+Size[random.randint(0, 4)]+Termination[random.randint(0, 1)]+Packaging[random.randint(0, 4)]+str(random.randint(10, 99))+ResistanceD1+str(random.randint(0, 9))+Tolerance[random.randint(0, 4)]+tcr[random.randint(0, 4)]
    
    faimly = 'RN73'
    zsize = Size[random.randint(0, 4)]
    zTermination=Termination[random.randint(0, 1)]
    zPackaging=Packaging[random.randint(0, 4)]
    zResistance = str(random.randint(10, 99))+ResistanceD1+str(random.randint(0, 9))
    zTolerance=Tolerance[random.randint(0, 4)]
    ztcr=tcr[random.randint(0, 4)]
    
    row = [zpart_number,faimly,zsize,zTermination,zPackaging,zResistance,zTolerance,ztcr]

    return row


def array_datarandom():
    randomlist = []
    for i in range(0,5):
        randomlist.append(datarandom())

    return randomlist

#print(array_datarandom())


#--------------------------------------------------------------------------


#wb = load_workbook("1.xlsx")
 # grab the active worksheet
wb = Workbook()
#load_workbook("1.xlsx")


ws = wb.active
ws.title = 'test'



for i in range(0,5):
    ws.append(rowrandom())

print(ws['C2'].value)

wb.save("genarated.xlsx")
